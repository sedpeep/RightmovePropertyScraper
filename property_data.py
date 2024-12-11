import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, TimeoutException
from selenium.webdriver.support import expected_conditions as EC
import re
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.wait import WebDriverWait


options = Options()
options.add_experimental_option("detach", True)
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
options.add_argument("--disable-blink-features=AutomationControlled")
options.add_argument("--disable-popup-blocking")
options.add_argument("--disable-save-password-bubble")


driver = webdriver.Chrome(options=options)
file_path = 'Rightmove scrape template.xlsx'

# Load the Excel file (ensure the file exists and is in .xlsx format)
try:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    print("Excel file loaded successfully.")
except FileNotFoundError:
    print(f"Error: The file at {file_path} was not found. Please check the path and try again.")
except openpyxl.utils.exceptions.InvalidFileException:
    print(f"Error: The file at {file_path} is not a valid Excel file. Please check the format.")

# Function to extract the postcode from the address
def extract_postcode(address):
    match = re.search(r'[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}', address)
    if match:
        return match.group(0)
    return ''

def check_tenanted(description_text):
    if "tenant" in description_text.lower() or "tenanted" in description_text.lower():
        return "Yes"
    return "No"

cookies_accepted = False

# Iterate through rows in the Excel sheet
for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, min_col=1, max_col=1):
    url_cell = row[0]
    url = url_cell.value

    if not url:
        continue  # Skip if URL is empty

    # Visit the URL
    driver.get(url)
    # Accept cookies only once
    if not cookies_accepted:
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="onetrust-accept-btn-handler"]'))
            ).click()
            cookies_accepted = True  # Mark cookies as accepted
        except (NoSuchElementException, TimeoutException):
            print("Cookies button not found or already clicked.")

    time.sleep(2)

    try:
        # Property Address
        try:
            property_address = driver.find_element(By.XPATH, '//*[@id="contact-agent-aside"]/div[1]/div/div[2]').text
        except NoSuchElementException:
            property_address = 'NULL'
        sheet.cell(row=url_cell.row, column=2).value = property_address

        # Postcode (from address)
        postcode = extract_postcode(property_address)
        sheet.cell(row=url_cell.row, column=3).value = postcode

        # Asking Price
        try:
            asking_price = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located(
                    (By.XPATH, '//*[@id="root"]/main/div/div[2]/div/article[1]/div/div/div[1]/span[1]')
                )
            ).text
        except (NoSuchElementException, TimeoutException):
            asking_price = 'NULL'
        sheet.cell(row=url_cell.row, column=4).value = asking_price

        # Property Type
        try:
            property_type = driver.find_element(By.XPATH, '/html/body/div[2]/main/div/div[2]/div/article[2]/dl/div[1]/dd/span/p').text
        except NoSuchElementException:
            property_type = 'NULL'
        sheet.cell(row=url_cell.row, column=5).value = property_type

        # Agent
        try:
            agent = driver.find_element(By.XPATH, '//*[@id="contact-agent-aside"]/div[1]/div/div[1]/a').text
        except NoSuchElementException:
            agent = 'NULL'
        sheet.cell(row=url_cell.row, column=6).value = agent

        # Size
        try:
            size = WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="info-reel"]/div[4]/dd/span/p[1]'))).text
        except (NoSuchElementException, TimeoutException):
            size = 'NULL'
        sheet.cell(row=url_cell.row, column=7).value = size

        # Bedrooms
        try:
            bedrooms = driver.find_element(By.XPATH, '//*[@id="info-reel"]/div[2]/dd/span/p').text
        except NoSuchElementException:
            bedrooms = 'NULL'
        sheet.cell(row=url_cell.row, column=8).value = bedrooms

        # Bathrooms
        try:
            bathrooms = driver.find_element(By.XPATH, '//*[@id="info-reel"]/div[3]/dd/span/p').text
        except NoSuchElementException:
            bathrooms = 'NULL'
        sheet.cell(row=url_cell.row, column=9).value = bathrooms

        # Leasehold or Freehold
        try:
            lease_status = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/main/div/div[2]/div/article[2]/dl/div[5]/dd/span/p'))
            ).text
        except (NoSuchElementException, TimeoutException):
            lease_status = 'NULL'
        sheet.cell(row=url_cell.row, column=10).value = lease_status

        # Lease Length if Leasehold
        if "leasehold" in lease_status.lower():
            try:
                button_element = driver.find_element(By.XPATH, '/html/body/div[2]/main/div/div[2]/div/article[3]/div[5]/button')
                driver.execute_script("arguments[0].scrollIntoView(true);", button_element)
                button_element.click()

                lease_length = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, '/html/body/div[2]/main/div/div[2]/div/article[3]/div[5]/div/div[2]/div/div[3]/p'))
                ).text
            except (NoSuchElementException, TimeoutException, ElementNotInteractableException):
                lease_length = 'NULL'
            sheet.cell(row=url_cell.row, column=11).value = lease_length
        else:
            sheet.cell(row=url_cell.row, column=11).value = 'Freehold'

        # Garden
        try:
            garden = driver.find_element(By.XPATH, '//*[@id="root"]/main/div/div[2]/div/article[3]/dl/div[3]/dd/span').text
        except NoSuchElementException:
            garden = 'NULL'
        sheet.cell(row=url_cell.row, column=12).value = garden

        # Parking
        try:
            parking = driver.find_element(By.XPATH, '//*[@id="root"]/main/div/div[2]/div/article[3]/dl/div[2]/dd/span').text
        except NoSuchElementException:
            parking = 'NULL'
        sheet.cell(row=url_cell.row, column=13).value = parking

        # Broadband Speed
        try:
            driver.find_element(By.XPATH, '//*[@id="root"]/main/div/div[2]/div/div[14]/button').click()
            broadband_speed = WebDriverWait(driver, 10).until(
                EC.visibility_of_element_located((By.XPATH, '/html/body/div[2]/main/div/div[2]/div/div[14]/div/div[2]/div[1]/div[1]/div/div/div/p[2]'))
            ).text
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException):
            broadband_speed = 'NULL'
        sheet.cell(row=url_cell.row, column=14).value = broadband_speed

        # EPC Rating
        try:
            epc_button = driver.find_element(By.XPATH, '//*[@id="epc-chevron"]/button')
            if epc_button.is_displayed():
                epc_button.click()
                epc_rating = driver.find_element(By.XPATH, '//*[@id="epc-chevron"]/div/div[2]/div[2]/div[1]/a').text
            else:
                epc_rating = 'NULL'
        except (NoSuchElementException, ElementNotInteractableException, TimeoutException):
            epc_rating = 'NULL'
        sheet.cell(row=url_cell.row, column=15).value = epc_rating

        # Council Tax Band
        try:
            council_tax_band = driver.find_element(By.XPATH, '//*[@id="root"]/main/div/div[2]/div/article[3]/dl/div[1]/dd').text
        except NoSuchElementException:
            council_tax_band = 'NULL'
        sheet.cell(row=url_cell.row, column=16).value = council_tax_band

        # Tenanted
        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="root"]/main/div/div[2]/div/article[3]/div[3]/button[2]'))
            ).click()
            full_description = ""
            paragraph_index = 1
            while True:
                try:
                    paragraph = driver.find_element(By.XPATH, f'//*[@id="root"]/main/div/div[2]/div/article[3]/div[3]/div/div[{paragraph_index}]').text
                    full_description += paragraph + " "
                    paragraph_index += 1
                except NoSuchElementException:
                    break
            tenanted_status = check_tenanted(full_description)
        except (NoSuchElementException, TimeoutException):
            tenanted_status = "No"
        sheet.cell(row=url_cell.row, column=17).value = tenanted_status

    except NoSuchElementException as e:
        print(f"Error extracting data for URL {url}: {e}")
        continue

# Save the Excel file
workbook.save('Rightmove scrape template.xlsx')

driver.quit()